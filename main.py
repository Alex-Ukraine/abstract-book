from collections import Counter
from copy import copy
from typing import List

from openpyxl import load_workbook
from pdfminer.high_level import extract_pages
from pdfminer.layout import LTTextBoxHorizontal, LTChar

file_pdf = "Abstract Book from the 5th World Psoriasis and Psoriatic Arthritis Conference 2018.pdf"
file_xml = "Data Entry - 5th World Psoriasis & Psoriatic Arthritis Conference 2018 - Case format (2).xlsx"
START_PAGE = 100
END_PAGE = 155
SHEET_NAME = "Sheet1"
START_ROW = 7


def update_xml(file: str, rec: list, sheet_name: str, start_row: int) -> None:
    """
    function load's xml file and then put's list of rows 'rec', start's with 'start_row'
    :param file: file xml for read and update with bunch of records
    :param rec: list of tuples aka records, we expect that every tuple consists 6 values
    :param sheet_name: name of sheet that has to be updated
    :param start_row: number of row from which we start to update cells
    """
    wb = load_workbook(file)
    ws = wb[sheet_name]
    for row in range(start_row, len(rec) + start_row):
        for col in range(1, 7):
            ws.cell(row=row, column=col).value = rec[row - start_row][col - 1]
            cell = ws["A3"]
            if cell.has_style:
                ws.cell(row=row, column=col).font = copy(cell.font)
                ws.cell(row=row, column=col).border = copy(cell.border)
                ws.cell(row=row, column=col).fill = copy(cell.fill)
                ws.cell(row=row, column=col).number_format = copy(cell.number_format)
                ws.cell(row=row, column=col).protection = copy(cell.protection)
                ws.cell(row=row, column=col).alignment = copy(cell.alignment)

    wb.save(file)
    wb.close()


def parse_articles(file: str, start_page: int, end_page: int) -> List[tuple]:
    """
    function parses pdf-file and return's list of tuples with received data
    :param file: pdf-file which we want to parse
    :param start_page: number of page where we start to parse
    :param end_page: number of page where we end to parse
    :return: list of tuples(ordered values)
    """
    # flag for breaking a loop if we reach END_PAGE
    flag = False

    # initial values
    session_name = ""
    topic_title = ""
    name = ""
    affiliation_name = ""
    persons_location = ""
    abstract = ""
    records = []

    # walking through pages
    for page_layout in extract_pages(file):
        if flag:
            break
        # walking through pdf basic elements
        for element in page_layout:
            if isinstance(element, LTTextBoxHorizontal):

                for text_line in element:
                    # let's do some investigation about most frequent font and size of one pdf-line
                    font_and_size_list = [
                        (character.fontname, character.size)
                        for character in text_line
                        if isinstance(character, LTChar)
                    ]
                    font_and_size = Counter(font_and_size_list).most_common(1)[0][0]

                    """
                    After my research I got interesting correlation:
                    ("OAOVWE+TimesNewRomanPS-BoldItalicMT", 9.5) correlate with session_name
                    ("OAOVWE+TimesNewRomanPS-BoldMT", 9.0) correlate with topic_title
                    ("OAOVWE+TimesNewRomanPS-ItalicMT", 9.0) correlate with name
                    ("OAOVWE+TimesNewRomanPS-ItalicMT", 8.0) correlate with affiliation_name
                    """

                    if font_and_size == ("OAOVWE+TimesNewRomanPS-BoldItalicMT", 9.5):
                        base_name = text_line.get_text()
                        if affiliation_name:
                            # we assume that every institutions text-box end's with location,
                            # ... but for most cases it's wrong idea
                            persons_location = ", ".join(
                                affiliation_name.split(", ")[-2:]
                            )
                            if "University" not in persons_location:
                                affiliation_name = affiliation_name.replace(
                                    ", " + persons_location, ""
                                )
                            else:
                                persons_location = ""

                        current_page = int(base_name[1:])
                        """if current_page < start_page:
                            break"""

                        if start_page < current_page <= end_page:
                            names = name.split(", ")
                            if names:
                                for _ in names:
                                    record = (
                                        _.replace("-\n", "").replace("\n", ""),
                                        affiliation_name.replace("-\n", "").replace(
                                            "\n", ""
                                        ),
                                        persons_location.replace("-\n", "").replace(
                                            "\n", ""
                                        ),
                                        session_name.replace("\n", ""),
                                        topic_title.replace("-\n", "").replace(
                                            "\n", ""
                                        ),
                                        abstract.replace("-\n", "").replace("\n", ""),
                                    )
                                    records.append(record)
                        # when we encounter with page more than end_page,
                        # ... we need to append accumulated temporary data to results
                        elif current_page > end_page:
                            names = name.split(", ")
                            if names:
                                for _ in names:
                                    record = (
                                        _.replace("-\n", "").replace("\n", ""),
                                        affiliation_name.replace("-\n", "").replace(
                                            "\n", ""
                                        ),
                                        persons_location.replace("-\n", "").replace(
                                            "\n", ""
                                        ),
                                        session_name.replace("\n", ""),
                                        topic_title.replace("-\n", "").replace(
                                            "\n", ""
                                        ),
                                        abstract.replace("-\n", "").replace("\n", ""),
                                    )
                                    records.append(record)
                            flag = True
                            break

                        # we need to remember last saved number of article
                        session_name = text_line.get_text()

                        # we need to empty temporary accumulators
                        topic_title = ""
                        name = ""
                        affiliation_name = ""
                        persons_location = ""
                        abstract = ""
                    elif font_and_size == ("OAOVWE+TimesNewRomanPS-BoldMT", 9.0):
                        # lets filter line from super-scripted numbers of references
                        line = "".join(
                            _.get_text()
                            for _ in text_line
                            if isinstance(_, LTChar) and _.size == 9.0
                        )
                        topic_title += line
                        # topic_title += text_line.get_text()
                    elif font_and_size == ("OAOVWE+TimesNewRomanPS-ItalicMT", 9.0):
                        line = "".join(
                            _.get_text()
                            for _ in text_line
                            if isinstance(_, LTChar) and _.size == 9.0
                        )
                        # check for absence 'References:' or 'Reference:' text-line and only then add line
                        if ":" not in line:
                            name += line
                    elif font_and_size == ("OAOVWE+TimesNewRomanPS-ItalicMT", 8.0):
                        line = "".join(
                            _.get_text()
                            for _ in text_line
                            if isinstance(_, LTChar) and _.size == 8.0
                        )
                        if ":" not in line:
                            affiliation_name += line
                    else:
                        abstract += text_line.get_text()
    return records


if __name__ == "__main__":
    update_xml(
        file_xml, parse_articles(file_pdf, START_PAGE, END_PAGE), SHEET_NAME, START_ROW
    )
